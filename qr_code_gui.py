
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.pagebreak import Break
import qrcode
from openpyxl.drawing.image import Image
from io import BytesIO
import tkinter as tk
from tkinter import messagebox
from openpyxl.utils import get_column_letter

def lancer_traitement(nom_fichier):
    try:
        nom = nom_fichier.strip()
        fichier_excel = nom + ".xlsx"
        if not os.path.exists(fichier_excel):
            raise FileNotFoundError(f"❌ Le fichier '{fichier_excel}' est introuvable.")

        match = re.search(r'(\d{6})', os.path.basename(fichier_excel))
        if not match:
            raise ValueError("❌ Aucun motif de date à 6 chiffres (JJMMAA) trouvé dans le nom du fichier.")
        date_brute = match.group(1)
        date_formatee = f"{date_brute[:2]}/{date_brute[2:4]}/{date_brute[4:]}"

        feuille_source = "Outbound"
        feuille_liste_cmd = "liste cmd"
        feuille_picking = "picking"

        df_source = pd.read_excel(fichier_excel, sheet_name=feuille_source)
        df_source.columns = df_source.columns.str.strip().str.lower()

        df_grouped = df_source.groupby('code commande', as_index=False).agg({
            'priorité': 'first',
            'quantité': 'sum'
        })
        colonnes_finales = ['code commande', 'priorité', 'quantité', 'date', 'prepa', 'ph debut', 'ph de fin',
                            'contrôle', 'c h debut', 'c h de fin']
        for col in colonnes_finales:
            if col not in df_grouped.columns:
                df_grouped[col] = ''
        df_grouped['date'] = date_formatee
        df_final = df_grouped[colonnes_finales]

        df_picking = df_source[['code commande', 'propriétaire commande', 'code article', 'quantité', 'code lot requis']].copy()
        df_picking['conditionnement'] = ''
        df_picking['date'] = date_formatee
        df_picking = df_picking[['code commande', 'propriétaire commande', 'date', 'code article', 'quantité', 'code lot requis', 'conditionnement']]
        df_picking['code lot requis'] = df_picking['code lot requis'].astype(str).str.replace("UNIC_SIZE_PRICE", "", regex=False)
        df_picking['code article'] = df_picking['code article'].apply(lambda x: int(x) if pd.notna(x) else x)
        df_picking['code article'] = df_picking['code article'].apply(lambda x: ' '.join(str(x)[i:i+3] for i in range(0, len(str(x)), 3)) if pd.notna(x) else '')
        df_picking.sort_values(by=['code commande', 'code lot requis'], inplace=True)

        with pd.ExcelWriter(fichier_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_final.to_excel(writer, sheet_name=feuille_liste_cmd, index=False)
            df_picking.to_excel(writer, sheet_name=feuille_picking, index=False)

        def appliquer_mise_en_forme(ws):
            font = Font(size=16)
            border = Border(left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin"))
            header_fill = PatternFill(start_color="B7CE9E", end_color="B7CE9E", fill_type="solid")
            alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            max_row = ws.max_row
            max_col = ws.max_column

            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                for cell in row:
                    cell.font = font
                    cell.border = border
                    if cell.row == 1:
                        cell.font = Font(size=16, bold=True, color="FFFFFF")
                        cell.fill = header_fill
                    elif cell.row % 2 == 0:
                        cell.fill = alt_fill

            col_a_values = [str(cell.value) for cell in ws['A'][1:] if cell.value is not None]
            max_len = max((len(v) for v in col_a_values), default=10)
            ws.column_dimensions['A'].width = max_len + 4

            for col in ws.iter_cols(min_col=2, max_col=max_col):
                col_letter = col[0].column_letter
                col_name = ws.cell(row=1, column=col[0].column).value
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)

                if col_letter == 'D':
                    ws.column_dimensions[col_letter].width = 35
                elif col_letter == 'B':
                    ws.column_dimensions[col_letter].width = 20
                elif col_letter == 'F':
                    ws.column_dimensions[col_letter].width = 30
                elif col_letter == 'G':
                    ws.column_dimensions[col_letter].width = 25
                elif col_name and col_name.strip().lower() in ['prepa', 'ph debut', 'ph de fin', 'contrôle', 'c h debut', 'c h de fin']:
                    ws.column_dimensions[col_letter].width = 15
                else:
                    ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                for cell in row:
                    if cell.row == 1:
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    else:
                        if cell.column_letter in ['A', 'F']:
                            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                        else:
                            cell.alignment = Alignment(wrap_text=False, horizontal='center', vertical='center')

            for r in range(1, max_row + 1):
                ws.row_dimensions[r].height = 40

        wb = load_workbook(fichier_excel)
        for feuille in [feuille_liste_cmd, feuille_picking]:
            ws = wb[feuille]
            appliquer_mise_en_forme(ws)

        ws = wb[feuille_picking]
        ws.print_title_rows = '1:1'
        ws.print_area = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
        ws.oddFooter.center.text = "Page &P / &N"
        ws.oddFooter.center.size = 12
        ws.oddFooter.center.font = "Arial"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = False
        ws.page_setup.fitToWidth = 1

        previous = None
        for row in range(2, ws.max_row + 1):
            current = ws.cell(row=row, column=1).value
            if previous is not None and current != previous:
                ws.row_breaks.append(Break(id=row - 1))
            previous = current

        highlight_fill = PatternFill(start_color="FFF176", end_color="FFF176", fill_type="solid")
        ws_cmd = wb[feuille_liste_cmd]
        header = [cell.value for cell in ws_cmd[1]]
        if "priorité" in header:
            col_index = header.index("priorité") + 1
            for row in range(2, ws_cmd.max_row + 1):
                cell = ws_cmd.cell(row=row, column=col_index)
                if str(cell.value).strip().lower() == "urgent":
                    cell.fill = highlight_fill

        df_code_brut = df_source[['code commande', 'code article']].copy()
        df_code_brut['code article'] = df_code_brut['code article'].apply(lambda x: int(x) if pd.notna(x) else x)

        col_art = [cell.value for cell in ws[1]].index("code article") + 1
        ws.insert_cols(col_art)
        ws.cell(row=1, column=col_art).value = "QR Code gauche"
        ws.insert_cols(col_art + 2)
        ws.cell(row=1, column=col_art + 2).value = "QR Code droite"
        appliquer_mise_en_forme(ws)

        header = [cell.value for cell in ws[1]]
        col_cmd = header.index("code commande") + 1
        col_art = header.index("code article") + 1
        col_qr_left = header.index("QR Code gauche") + 1
        col_qr_right = header.index("QR Code droite") + 1

        ws.column_dimensions[get_column_letter(col_qr_left)].width = 18
        ws.column_dimensions[get_column_letter(col_qr_right)].width = 18

        for row in range(2, ws.max_row + 1):
            code_commande = ws.cell(row=row, column=col_cmd).value
            code_article_formate = ws.cell(row=row, column=col_art).value

            if code_commande and code_article_formate:
                code_article_nettoye = code_article_formate.replace(" ", "")
                match = df_source[
                    (df_source['code commande'] == code_commande) &
                    (df_source['code article'].astype(str).str.contains(code_article_nettoye))
                ]
                if not match.empty:
                    code_article_brut = match.iloc[0]['code article']
                    if pd.notna(code_article_brut):
                        qr = qrcode.make(str(int(code_article_brut)))
                        img_io = BytesIO()
                        qr.save(img_io, format='PNG')
                        img_io.seek(0)
                        img = Image(img_io)
                        img.width = img.height = 60

                        if (row - 2) % 2 == 0:
                            ws.add_image(img, ws.cell(row=row, column=col_qr_left).coordinate)
                        else:
                            ws.add_image(img, ws.cell(row=row, column=col_qr_right).coordinate)

                        ws.row_dimensions[row].height = 70

        wb.save(fichier_excel)
        messagebox.showinfo("Succès", "✅ Fichier mis en page avec succès.")

    except Exception as e:
        messagebox.showerror("Erreur", f"❌ Une erreur s'est produite :\n{str(e)}")

root = tk.Tk()
root.title("Traitement Picking QR Code")
root.geometry("400x150")

label = tk.Label(root, text="Nom du fichier (sans extension) :", font=("Arial", 12))
label.pack(pady=10)

entry = tk.Entry(root, font=("Arial", 12))
entry.pack()

btn = tk.Button(root, text="Lancer", font=("Arial", 12), bg="#4CAF50", fg="white",
                command=lambda: lancer_traitement(entry.get()))
btn.pack(pady=20)

root.mainloop()
